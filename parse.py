import csv
from email.policy import default
from sys import argv
from tkinter.tix import ROW
from traceback import print_exc
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl.styles.colors as xlcolour
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter


HEIGHT      = 4 # events start at 4 different times in hour
DAY_WIDTH   = 4
DAY_HEIGHT  = 60
FILE_NAME   = "out1.xlsx"

ROW_BASE = 3
COLUMN_BASE = 3

class Event:
    def __init__(self, row: list[str]) -> None:
        (self.name,
        self.group,
        self.type,
        self.title,
        self.info,
        self.wday,
        self.first_day,
        self.last_day,
        self.start_time,
        self.end_time,
        self.place,
        self.capacity,
        self.teacher,
        self.email,
        self.required_services,
        self.accepted,
        self.artefact) = row

    
    def cell_column(self) -> tuple[int, int]:
        """
        Generate column range for excel
        return (
            start_column,
            end_column
        )
        """
        # First generate width  -> if LECTURE   width 6
        #                       -> if CWA       width 2
        #                       -> if CWL       width 1

        width = {
            "Wykład": DAY_WIDTH,
            "CWA":  DAY_WIDTH //2,
            "CWL":  DAY_WIDTH // 4
        }.get(self.type, 1)

        cell = {
            "Pn": 0,
            "Wt": 1,
            "Śr": 2,
            "Cz": 3,
            "Pt": 4
        }.get(self.wday, None) * DAY_WIDTH + 1 + (int(self.group.strip("a")) - 1) * width

        return int(cell) + COLUMN_BASE, int(cell + width - 1) + COLUMN_BASE

    def cell_row(self) -> int:
        """Generate row range for excel
        return (
            row_start,
            row_end
        )"""
        h, m = self.start_time.split(":")
        h = int(h)
        m = int(m)

        start = h * 4 + m / 15 - 31 + ROW_BASE

        h, m = self.end_time.split(":")
        h = int(h)
        m = int(m)

        end = h * 4 + m / 15 - 31 + ROW_BASE

        return int(start), int(end) - 1

    def colour(self):
        return {
            "Wykład": "CC6633",
            "CWA": "33FF33",
            "CWL": "5555FF"
        }.get(self.type, "333333")

    def value(self) -> str:
        title = self.title.strip("(2 rok)")
        title = ". ".join([x[:4] for x in title.split(" ")])
        teacher = ", ".join(filter(lambda x: len(x)>2, [t.split(" ")[-1] for t in self.teacher.split(",")]))
        type_ = self.type if self.type != "Wykład" else "W" 
        return ", ".join([ type_, title, self.place, teacher, self.start_time, self.end_time])


def set_day_borders(sh: Worksheet):
    thick = Side(style="thick")

    # TOP
    top_border = Border(top=thick)
    for cell in cell_range(sh, ROW_BASE + 1, ROW_BASE + 1, COLUMN_BASE + 1, COLUMN_BASE + DAY_WIDTH * 5):
        cell.border = top_border

    # LEFT
    left_border = Border(left=thick)
    for cell in cell_range(sh, ROW_BASE + 1, ROW_BASE + DAY_HEIGHT, COLUMN_BASE + 1, COLUMN_BASE + 1):
        cell.border = left_border

    # RIGHT
    right_border = Border(right=thick)
    for cell in cell_range(sh, ROW_BASE + 1, ROW_BASE + DAY_HEIGHT, COLUMN_BASE + DAY_WIDTH * 5, COLUMN_BASE + DAY_WIDTH * 5):
        cell.border = right_border

    # Fix corners
    # left top
    sh.cell(row=ROW_BASE + 1, column=COLUMN_BASE + 1).border = Border(top=thick, left=thick)
    
    # right top
    sh.cell(row=ROW_BASE + 1, column=COLUMN_BASE + DAY_WIDTH * 5).border = Border(top=thick, right=thick)

    # thick border between days
    for i in range(1, 5):
        col = COLUMN_BASE + 1 + i * DAY_WIDTH
        sh.cell(ROW_BASE + 1, col).border = Border(left=thick, top=thick)
        for cell in cell_range(sh, ROW_BASE + 1 + 1, ROW_BASE + DAY_HEIGHT, col, col):
            cell.border = left_border


def sheet_modify_width_height(sh: Worksheet):
    for i in range(COLUMN_BASE + 1, COLUMN_BASE + 1 + DAY_WIDTH * 5):
        sh.column_dimensions[get_column_letter(i)].width *= 1.02

    for i in range(ROW_BASE + 1, ROW_BASE + DAY_HEIGHT):
        sh.row_dimensions[i].height = 10

def cell_range(worksheet: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        for row in range(start_row, end_row + 1):
            yield worksheet.cell(column=col, row=row)



with open(argv[1], 'r', encoding='utf-8') as f:
    events = csv.reader(f)

    for row in events:
        for i, x in enumerate(row):
            print(i, x)
        break

    xl = Workbook()
    sh = xl.active

    set_day_borders(sh)
    sheet_modify_width_height(sh)

    thin_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))
            
    for i, event in enumerate(events):
        if i == 0:
            print(event)
            continue
        try:
            e = Event(event)
            if e.type == "Lektorat":
                print("continue lektorat:", event)
                continue

            print("happeinng")
            start_col, end_col = e.cell_column()
            start_row, end_row = e.cell_row()


            sh.merge_cells(start_row=start_row, end_row=end_row, start_column=start_col, end_column=end_col)
            for cell in cell_range(sh, start_row, end_row, start_col, end_col):
                cell.border = thin_border
            
            
            c = sh.cell(row=start_row, column=start_col)
            print(c)

            c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            c.fill = PatternFill(patternType='solid',  fgColor=xlcolour.Color(rgb="00"+e.colour()))
            c.value = e.value()
            c.font = Font(size=8)
            print("Success: ", i)

        except:
            print_exc()
            print(i, event)
            xl.save(FILE_NAME)
            exit()

    xl.save(".".join(argv[1].split(".")[:-1]) + ".xlsx")



        




